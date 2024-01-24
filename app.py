# begin with pip install openpyxl on the terminal (bottom left hand menu).
# this needs to be done per project!
import random

import openpyxl as xl
import random as ran

wb = xl.load_workbook('colours.xlsx')
colours = wb['Sheet1']
define = wb['Sheet2']

# # cell = define.cell(1, 1)
# # print(cell.value)
#
# # print(colours.max_column)
#
# print(define)
#
# for column in range(1, colours.max_column+1):
#     # print(colours.cell(1, column).value)
#     var_name = colours.cell(1, column).value
#     locals()[var_name] = ""
#
#     for row in range(2, colours.max_row+1):
#         if colours.cell(row, column).value is None:
#             pass
#         else:
#             # print(
#             #     colours.cell(row, column).value
#             #     + str(colours.cell(row, column))
#             # )
#             locals()[var_name] += colours.cell(row, column).value  # This needs to add to list not concat
#
#
# print(Blue)
# print(Red)
# print(Green)
#
# # We don't actually need to query the var names.
# # We need to randomly pick a column then pick two numbers from inside it
# # Could cell references do this?

maxcol = colours.max_column

# print(maxcol)

# print(random.randint(1, maxcol))

colselect = ran.randint(1, maxcol)

# print(colselect)

selection = []
header = colours.cell(1, colselect).value

for rows in range(2, colours.max_row+1):
    if colours.cell(rows, colselect).value is None:
        pass
    else:
        # print(colours.cell(rows, colselect).value)
        value = colours.cell(rows, colselect).value
        selection.append(value)

# print(selection)

firstcol = ran.choice(selection)
secondcol = ran.choice(selection)

# print(firstcol+secondcol)

if firstcol == secondcol:
    print(f'{header}: {firstcol} only')
else:
    print(f'{header}: {firstcol} and {secondcol}')

