import openpyxl as xl
import random as ran

wb = xl.load_workbook('colours2.xlsx')

print(wb.sheetnames)

colours = wb['Sheet1']

max_col = colours.max_column

# print(max_col)

# print(random.randint(1, max_col))

col_select = ran.randint(1, max_col)

# print(col_select)

selection = []
header = colours.cell(1, col_select).value

for rows in range(2, colours.max_row+1):
    if colours.cell(rows, col_select).value is None:
        pass
    else:
        # print(colours.cell(rows, col_select).value)
        value = colours.cell(rows, col_select).value
        selection.append(value)

# print(selection)

first_col = ran.choice(selection)
second_col = ran.choice(selection)

# print(firstcol+secondcol)

if first_col == second_col:
    print(f'{header}: {first_col} only')
else:
    print(f'{header}: {first_col} and {second_col}')

