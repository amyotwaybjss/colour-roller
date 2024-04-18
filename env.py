# this sheet will create our 'settings', load in sheet, get sheet list, store vars

import openpyxl as xl
import os.path as path
import inspect
from data_clean import data_clean
from user_settings import wb_name, force_refresh


# Creating required elements
# inspect.cleandoc is cleaning up the indentation

new_line = "\n"

introduction = inspect.cleandoc('''
    Hello, Welcome! 
    Select a mode or press enter to skip.
    ''')

instructions = new_line + inspect.cleandoc('''
        sheet name = Rolls from that sheet.
        r = Re-Rolls using current settings.
        config = Adjusts randomness settings.
        mode = Returns to the mode select screen.
        quit = Quits the program.
        (Type 'help' to repeat this message)
        ''') + new_line

shuffle_detail = new_line + inspect.cleandoc('''
        Shuffle Settings:
        mix = Picks from two different lists
        match = Picks from same list
        random = Picks either from the same list or different list
        ''') + new_line

# Creating required functions
# Note that without a function, importing variables does not work correctly


def clean_exists(load, name):
    if path.isfile(f'./{name}_cleaned.xlsx') is False or force_refresh is True:
        data_clean(load, name)


def status(mode, sheets):
    stat = f'Mode: {mode}{new_line}Active Sheets: {sheets}'
    return stat
# as instructions were not dynamically updating the status func was moved out


# Checks to see if file is available

try:
    wb_input = xl.load_workbook(f'{wb_name}.xlsx')
except FileNotFoundError:
    print('Error: Named sheet has not been found, please check the file and re-run.')
    quit()
except PermissionError:
    print('Error: Unable to access file, please check file is closed and re-run.')
    quit()

# Checks if the clean version of the sheet exists, and if not creates one and loads it

clean_exists(wb_input, wb_name)
wb = xl.load_workbook(f'{wb_name}_cleaned.xlsx')
sheet_list = wb.sheetnames
